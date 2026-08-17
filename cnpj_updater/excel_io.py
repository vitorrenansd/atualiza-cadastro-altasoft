"""Leitura e escrita da planilha.

A entrada nunca e modificada: o export carrega o arquivo original, acrescenta
as colunas no fim e salva com outro nome. As colunas e formatacao existentes
sao preservadas porque o openpyxl reescreve a pasta de trabalho carregada, em
vez de montar uma nova.

Nada do que a planilha ja tem e sobrescrito. Quando a base antiga tem contato
proprio, as colunas de veredito dizem se ele confere com o da Receita, o que
transforma "nao sei se esses e-mails estao atualizados" numa coluna filtravel.
"""

import json
import re
from dataclasses import dataclass, field

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import cnpj as cnpj_util
from .config import Config
from .store import Store

# (cabecalho, chave no banco). Ordem = ordem das colunas acrescentadas.
COLUNAS_BASE = [
    ("Situação Cadastral", "situacao_cadastral"),
    ("Data Situação", "data_situacao"),
    ("CNAE Principal", "cnae_principal_codigo"),
    ("CNAE Principal — Descrição", "cnae_principal_descricao"),
    ("CNAEs Secundários", "_cnaes_secundarios"),
    ("Qtd CNAEs Sec.", "_qtd_secundarios"),
    ("Telefone 1 (Receita)", "telefone_1"),
    ("Telefone 2 (Receita)", "telefone_2"),
    ("E-mail (Receita)", "email"),
    ("Razão Social (Receita)", "razao_social"),
    ("Nome Fantasia (Receita)", "nome_fantasia"),
    ("Fonte", "fonte"),
    ("Status Consulta", "status"),
    ("Consultado em", "atualizado_em"),
]

COLUNAS_COMPARACAO = [
    ("Telefone confere?", "_cmp_telefone"),
    ("E-mail confere?", "_cmp_email"),
]

LARGURAS = {
    "Situação Cadastral": 18, "Data Situação": 14, "CNAE Principal": 15,
    "CNAE Principal — Descrição": 45, "CNAEs Secundários": 60,
    "Qtd CNAEs Sec.": 14, "Telefone 1 (Receita)": 20,
    "Telefone 2 (Receita)": 20, "E-mail (Receita)": 32,
    "Razão Social (Receita)": 40, "Nome Fantasia (Receita)": 30,
    "Fonte": 14, "Status Consulta": 16, "Consultado em": 22,
    "Telefone confere?": 18, "E-mail confere?": 18,
}

_FILL = PatternFill("solid", fgColor="DDEBF7")
_FILL_CMP = PatternFill("solid", fgColor="FCE4D6")
_FONT = Font(bold=True)


def colunas(cfg: Config) -> list[tuple[str, str]]:
    if cfg.comparar_contato:
        return COLUNAS_BASE + COLUNAS_COMPARACAO
    return list(COLUNAS_BASE)


@dataclass
class Linha:
    numero: int          # linha real na planilha (1-based)
    bruto: object        # valor como esta na celula
    cnpj: str | None     # normalizado para 14 digitos, ou None
    valido: bool
    email_atual: str = ""
    telefones_atuais: list[str] = field(default_factory=list)


# -- comparacao com o contato existente -----------------------------------

def _digitos(valor) -> str:
    return re.sub(r"\D", "", str(valor or ""))


def _veredito_email(atual: str, receita: str) -> str:
    a = (atual or "").strip().casefold()
    r = (receita or "").strip().casefold()
    if not a and not r:
        return "ambos vazios"
    if not r:
        return "só na planilha"
    if not a:
        return "só na Receita"
    return "igual" if a == r else "divergente"


def _veredito_email_com_status(atual: str, receita: str, status_email: str) -> str:
    """Veredito de e-mail, distinguindo "a Receita nao tem" de "nao perguntamos".

    A fase 1 e resolvida na maior parte por BrasilAPI/Minha Receita, que nao
    retornam e-mail. Nessas linhas, dizer "so na planilha" afirmaria que a
    Receita nao tem e-mail — conclusao errada, porque o campo nunca foi
    consultado. So a fase 2 autoriza esse veredito.
    """
    if status_email == "pendente" and not (receita or "").strip():
        return "não consultado"
    return _veredito_email(atual, receita)


def _veredito_telefone(atuais: list[str], *receita: str) -> str:
    def chaves(valores):
        # Compara pelos ultimos 8 digitos: tolera DDD ausente na base antiga e
        # o nono digito que os celulares ganharam depois. Comparar a string
        # inteira marcaria como divergente numero que e' o mesmo.
        saida = set()
        for v in valores:
            d = _digitos(v)
            if len(d) >= 8:
                saida.add(d[-8:])
        return saida

    da_planilha = chaves(atuais)
    da_receita = chaves(receita)
    if not da_planilha and not da_receita:
        return "ambos vazios"
    if not da_receita:
        return "só na planilha"
    if not da_planilha:
        return "só na Receita"
    return "igual" if da_planilha & da_receita else "divergente"


# -- leitura ---------------------------------------------------------------

def _escolher_aba(wb, nome: str):
    if nome:
        if nome not in wb.sheetnames:
            raise SystemExit(
                f"Aba '{nome}' nao existe. Abas disponiveis: {wb.sheetnames}"
            )
        return wb[nome]
    return wb[wb.sheetnames[0]]


def _mapa_cabecalhos(ws, linha_cabecalho: int) -> dict[str, int]:
    mapa = {}
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=linha_cabecalho, column=col).value
        if valor is not None:
            mapa[str(valor).strip().casefold()] = col
    return mapa


def _achar_coluna(ws, linha_cabecalho: int, nome_desejado: str) -> int:
    """Localiza a coluna de CNPJ. Tres estrategias, da mais explicita a menos."""
    cabecalhos = {}
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=linha_cabecalho, column=col).value
        if valor is not None:
            cabecalhos[col] = str(valor).strip()

    # 1) Nome exato pedido no config.
    if nome_desejado:
        alvo = nome_desejado.strip().casefold()
        for col, texto in cabecalhos.items():
            if texto.casefold() == alvo:
                return col
        raise SystemExit(
            f"Coluna '{nome_desejado}' nao encontrada na linha "
            f"{linha_cabecalho}. Cabecalhos lidos: {list(cabecalhos.values())}"
        )

    # 2) Cabecalho que mencione CNPJ ou inscricao.
    for chave in ("cnpj", "inscri"):
        for col, texto in cabecalhos.items():
            if chave in texto.casefold():
                return col

    # 3) Ultimo recurso: a coluna cujos valores mais parecem CNPJ. Protege
    #    contra planilha sem cabecalho ou com cabecalho inesperado.
    melhor, melhor_taxa = None, 0.0
    limite = min(ws.max_row, linha_cabecalho + 200)
    for col in range(1, ws.max_column + 1):
        total = acertos = 0
        for row in range(linha_cabecalho + 1, limite + 1):
            valor = ws.cell(row=row, column=col).value
            if valor is None or str(valor).strip() == "":
                continue
            total += 1
            norm = cnpj_util.normalizar(valor)
            if norm and cnpj_util.valido(norm):
                acertos += 1
        if total >= 5:
            taxa = acertos / total
            if taxa > melhor_taxa:
                melhor, melhor_taxa = col, taxa
    if melhor and melhor_taxa >= 0.6:
        print(f"  coluna de CNPJ detectada automaticamente: "
              f"{get_column_letter(melhor)} ({melhor_taxa:.0%} dos valores validam)")
        return melhor

    raise SystemExit(
        "Nao consegui identificar a coluna de CNPJ. Defina 'coluna_cnpj' "
        f"no config.toml. Cabecalhos lidos: {list(cabecalhos.values())}"
    )


def ler_linhas(cfg: Config) -> tuple[list[Linha], str, int]:
    """Le os CNPJs (e o contato atual) da planilha."""
    if not cfg.entrada.exists():
        raise SystemExit(f"Planilha de entrada nao encontrada: {cfg.entrada}")
    # data_only=True para pegar o valor calculado, caso o CNPJ venha de formula.
    wb = load_workbook(cfg.entrada, data_only=True)
    try:
        ws = _escolher_aba(wb, cfg.aba)
        col = _achar_coluna(ws, cfg.linha_cabecalho, cfg.coluna_cnpj)

        col_email = None
        cols_tel: list[int] = []
        if cfg.comparar_contato:
            mapa = _mapa_cabecalhos(ws, cfg.linha_cabecalho)
            if cfg.coluna_email_atual:
                col_email = mapa.get(cfg.coluna_email_atual.strip().casefold())
                if col_email is None:
                    print(f"  aviso: coluna de e-mail atual "
                          f"'{cfg.coluna_email_atual}' nao encontrada")
            for nome in cfg.colunas_telefone_atuais:
                idx = mapa.get(nome.strip().casefold())
                if idx is None:
                    print(f"  aviso: coluna de telefone '{nome}' nao encontrada")
                else:
                    cols_tel.append(idx)

        linhas: list[Linha] = []
        vazias_seguidas = 0
        for row in range(cfg.linha_cabecalho + 1, ws.max_row + 1):
            bruto = ws.cell(row=row, column=col).value
            if bruto is None or _digitos(bruto) == "":
                vazias_seguidas += 1
                # max_row costuma vir inflado por formatacao residual.
                if vazias_seguidas >= 50:
                    break
                continue
            vazias_seguidas = 0
            norm = cnpj_util.normalizar(bruto)
            linhas.append(Linha(
                numero=row,
                bruto=bruto,
                cnpj=norm,
                valido=bool(norm) and cnpj_util.valido(norm),
                email_atual=(
                    str(ws.cell(row=row, column=col_email).value or "").strip()
                    if col_email else ""
                ),
                telefones_atuais=[
                    str(ws.cell(row=row, column=c).value or "") for c in cols_tel
                ],
            ))
        return linhas, ws.title, col
    finally:
        wb.close()


# -- escrita ---------------------------------------------------------------

def exportar(cfg: Config, store: Store) -> dict:
    """Escreve a planilha de saida com as colunas acrescentadas."""
    linhas, _, _ = ler_linhas(cfg)
    dados = store.buscar_todos()
    layout = colunas(cfg)

    wb = load_workbook(cfg.entrada)  # sem data_only: preserva formulas
    ws = _escolher_aba(wb, cfg.aba)

    # Se o export ja rodou sobre este arquivo, reescreve as mesmas colunas em
    # vez de acrescentar um segundo bloco identico.
    existentes = {
        str(ws.cell(row=cfg.linha_cabecalho, column=c).value or "").strip(): c
        for c in range(1, ws.max_column + 1)
    }
    ja_nossas = [existentes[t] for t, _ in layout if t in existentes]
    inicio = min(ja_nossas) if ja_nossas else ws.max_column + 1

    for i, (titulo, chave) in enumerate(layout):
        celula = ws.cell(row=cfg.linha_cabecalho, column=inicio + i, value=titulo)
        celula.font = _FONT
        celula.fill = _FILL_CMP if chave.startswith("_cmp_") else _FILL
        celula.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(inicio + i)].width = LARGURAS.get(
            titulo, 18
        )

    contagem = {
        "preenchidas": 0, "sem_dados": 0, "invalidas": 0, "ativas": 0,
        "tel_divergente": 0, "tel_so_receita": 0,
        "email_divergente": 0, "email_so_receita": 0,
    }
    ativas_rows = []

    for linha in linhas:
        registro = dados.get(linha.cnpj) if linha.cnpj else None
        valores = _montar_valores(linha, registro, layout)
        for i, valor in enumerate(valores):
            ws.cell(row=linha.numero, column=inicio + i, value=valor)

        if not linha.valido:
            contagem["invalidas"] += 1
        elif registro and registro["status"] == "ok":
            contagem["preenchidas"] += 1
            if registro["situacao_cadastral"] == "Ativa":
                contagem["ativas"] += 1
                ativas_rows.append(linha.numero)
            if cfg.comparar_contato:
                v_tel = _veredito_telefone(
                    linha.telefones_atuais,
                    registro["telefone_1"], registro["telefone_2"],
                )
                v_mail = _veredito_email_com_status(
                    linha.email_atual, registro["email"],
                    registro["status_email"],
                )
                if v_tel == "divergente":
                    contagem["tel_divergente"] += 1
                elif v_tel == "só na Receita":
                    contagem["tel_so_receita"] += 1
                if v_mail == "divergente":
                    contagem["email_divergente"] += 1
                elif v_mail == "só na Receita":
                    contagem["email_so_receita"] += 1
        else:
            contagem["sem_dados"] += 1

    if cfg.aba_somente_ativas:
        _escrever_aba_ativas(wb, ws, cfg, ativas_rows, inicio + len(layout) - 1)

    cfg.saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(cfg.saida)
    wb.close()
    contagem["total_linhas"] = len(linhas)
    return contagem


def _montar_valores(linha: Linha, registro, layout) -> list:
    chaves = [c for _, c in layout]

    def so_status(mensagem: str) -> list:
        # A mensagem vai na coluna 'Status Consulta' pela chave, nao pela
        # posicao: o bloco de colunas muda de tamanho conforme
        # comparar_contato, e escrever "na ultima" acerta a coluna errada.
        saida = [""] * len(chaves)
        if "status" in chaves:
            saida[chaves.index("status")] = mensagem
        return saida

    if not linha.valido:
        return so_status("CNPJ invalido na planilha")
    if registro is None:
        return so_status("nao importado")

    secundarios = json.loads(registro["cnaes_secundarios"] or "[]")
    texto_sec = "; ".join(f"{c} {d}".strip() for c, d in secundarios)
    tem_dados = registro["status"] == "ok"

    saida = []
    for _, chave in layout:
        if chave == "_cnaes_secundarios":
            saida.append(texto_sec)
        elif chave == "_qtd_secundarios":
            saida.append(len(secundarios) if tem_dados else "")
        elif chave == "_cmp_telefone":
            saida.append(_veredito_telefone(
                linha.telefones_atuais,
                registro["telefone_1"], registro["telefone_2"],
            ) if tem_dados else "")
        elif chave == "_cmp_email":
            saida.append(_veredito_email_com_status(
                linha.email_atual, registro["email"], registro["status_email"]
            ) if tem_dados else "")
        else:
            saida.append(registro[chave] if registro[chave] is not None else "")
    return saida


def _escrever_aba_ativas(wb, ws_origem, cfg: Config, linhas_ativas: list[int],
                         ultima_col: int) -> None:
    """Aba extra so com as empresas ATIVAS (valores, sem formatacao)."""
    nome = "Somente Ativas"
    if nome in wb.sheetnames:
        del wb[nome]
    ws = wb.create_sheet(nome)

    for c in range(1, ultima_col + 1):
        celula = ws.cell(
            row=1, column=c,
            value=ws_origem.cell(row=cfg.linha_cabecalho, column=c).value,
        )
        celula.font = _FONT
        celula.fill = _FILL
    for destino, origem in enumerate(linhas_ativas, start=2):
        for c in range(1, ultima_col + 1):
            ws.cell(row=destino, column=c,
                    value=ws_origem.cell(row=origem, column=c).value)
    ws.freeze_panes = "A2"
